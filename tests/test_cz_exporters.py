# tests/test_cz_exporters.py
"""
Tests for CZ JSON and XLSX exporters.

Covers:
1. JSON contains summary + items + FTC data
2. JSON preserves exempt and pending items
3. JSON preserves standalone/unlinked WHT item
4. XLSX creates expected sheets
5. XLSX contains exempt security item
6. XLSX contains option item
7. XLSX contains FTC audit columns for dividend
8. XLSX standalone/unlinked WHT visible in WithholdingTax sheet
9. Exporters don't crash on empty/missing sections
10. JSON Decimal serialisation consistency
"""
import io
import json
import os
import tempfile
import uuid
from decimal import Decimal
from typing import List

import pytest

from src.countries.cz.config import CzTaxConfig
from src.countries.cz.enums import CzTaxSection
from src.countries.cz.exporters.json_exporter import export_cz_to_json
from src.countries.cz.exporters.xlsx_exporter import export_cz_to_xlsx
from src.countries.cz.fx_policy import CzFxPolicyConfig
from src.countries.cz.plugin import CzechTaxAggregator, CzechTaxClassifier
from src.countries.cz.tax_items import (
    CzExemptionReason,
    CzTaxItem,
    CzTaxItemType,
    CzTaxReviewStatus,
    CzWhtRecord,
)
from src.countries.base import TaxResult, TaxResultSection
from src.domain.enums import AssetCategory, FinancialEventType, RealizationType
from src.domain.events import CashFlowEvent, FinancialEvent, WithholdingTaxEvent
from src.domain.results import RealizedGainLoss
from src.identification.asset_resolver import AssetResolver
from src.classification.asset_classifier import AssetClassifier


# ---------------------------------------------------------------------------
# Fixtures — build a realistic TaxResult
# ---------------------------------------------------------------------------

def _resolver():
    class D(AssetClassifier):
        def __init__(self): super().__init__(cache_file_path="d.json")
        def save_classifications(self): pass
    return AssetResolver(asset_classifier=D())


def _make_rgl(gross, cat=AssetCategory.STOCK, holding_days=200, proceeds=Decimal("1500")):
    return RealizedGainLoss(
        originating_event_id=uuid.uuid4(),
        asset_internal_id=uuid.uuid4(),
        asset_category_at_realization=cat,
        acquisition_date="2024-06-15",
        realization_date="2025-03-25",
        realization_type=RealizationType.LONG_POSITION_SALE if cat != AssetCategory.OPTION
            else RealizationType.OPTION_TRADE_CLOSE_LONG,
        quantity_realized=Decimal("10"),
        unit_cost_basis_eur=Decimal("100"),
        unit_realization_value_eur=proceeds / Decimal("10"),
        total_cost_basis_eur=Decimal("1000"),
        total_realization_value_eur=proceeds,
        gross_gain_loss_eur=gross,
        holding_period_days=holding_days,
    )


def _build_test_result() -> TaxResult:
    """Build a realistic TaxResult with various item types for export testing."""
    resolver = _resolver()
    cfg = CzTaxConfig(annual_exempt_limit_enabled=False)

    rgls = [
        _make_rgl(Decimal("500"), AssetCategory.STOCK, holding_days=200),
        _make_rgl(Decimal("800"), AssetCategory.STOCK, holding_days=1200),  # exempt by time test
        _make_rgl(Decimal("300"), AssetCategory.OPTION, holding_days=100),
        _make_rgl(Decimal("-100"), AssetCategory.OPTION, holding_days=50),
    ]
    # Make one pending (no acquisition_date)
    rgl_pending = _make_rgl(Decimal("200"), AssetCategory.STOCK, holding_days=100)
    rgl_pending.acquisition_date = ""
    rgl_pending.holding_period_days = None
    rgls.append(rgl_pending)

    classifier = CzechTaxClassifier(config=cfg)
    for rgl in rgls:
        classifier.classify(rgl)

    div = CashFlowEvent(
        asset_internal_id=uuid.uuid4(),
        event_date="2025-06-15",
        event_type=FinancialEventType.DIVIDEND_CASH,
        gross_amount_foreign_currency=Decimal("100"),
        local_currency="USD",
        gross_amount_eur=Decimal("90"),
    )
    wht_linked = WithholdingTaxEvent(
        asset_internal_id=div.asset_internal_id,
        event_date="2025-06-15",
        gross_amount_foreign_currency=Decimal("15"),
        local_currency="USD",
        gross_amount_eur=Decimal("13.5"),
        taxed_income_event_id=div.event_id,
        source_country_code="US",
    )
    interest = CashFlowEvent(
        asset_internal_id=uuid.uuid4(),
        event_date="2025-06-15",
        event_type=FinancialEventType.INTEREST_RECEIVED,
        gross_amount_foreign_currency=Decimal("25"),
        local_currency="EUR",
        gross_amount_eur=Decimal("25"),
    )
    # Orphan WHT — different asset
    orphan_wht = WithholdingTaxEvent(
        asset_internal_id=uuid.uuid4(),
        event_date="2025-06-15",
        gross_amount_foreign_currency=Decimal("10"),
        local_currency="USD",
        gross_amount_eur=Decimal("9"),
        source_country_code="DE",
    )

    events: List[FinancialEvent] = [div, wht_linked, interest, orphan_wht]

    aggregator = CzechTaxAggregator(config=cfg)
    return aggregator.aggregate(rgls, events, resolver, 2025)


# =========================================================================
# JSON Tests
# =========================================================================

class TestJsonExport:
    def setup_method(self):
        self.result = _build_test_result()

    def test_json_contains_summary_items_ftc(self):
        json_str = export_cz_to_json(self.result)
        data = json.loads(json_str)

        assert "metadata" in data
        assert data["metadata"]["country"] == "cz"
        assert data["metadata"]["tax_year"] == 2025
        assert "sections" in data
        assert "items" in data
        assert len(data["items"]) > 0
        assert "ftc_summary" in data

    def test_json_preserves_exempt_items(self):
        json_str = export_cz_to_json(self.result)
        data = json.loads(json_str)

        exempt_items = [i for i in data["items"] if i.get("is_exempt") is True]
        assert len(exempt_items) >= 1
        assert any(i["exemption_reason"] == "TIME_TEST_PASSED" for i in exempt_items)

    def test_json_preserves_pending_items(self):
        json_str = export_cz_to_json(self.result)
        data = json.loads(json_str)

        pending = [i for i in data["items"] if i.get("tax_review_status") == "PENDING_MANUAL_REVIEW"]
        assert len(pending) >= 1
        assert data["warnings"]["pending_review_count"] >= 1
        assert len(data["warnings"]["pending_review_event_ids"]) >= 1

    def test_json_preserves_unlinked_wht(self):
        json_str = export_cz_to_json(self.result)
        data = json.loads(json_str)

        other_items = [i for i in data["items"] if i.get("item_type") == "OTHER"]
        assert len(other_items) >= 1
        assert data["warnings"]["unlinked_wht_count"] >= 1
        assert len(data["warnings"]["unlinked_wht_event_ids"]) >= 1

    def test_json_ftc_data_present(self):
        json_str = export_cz_to_json(self.result)
        data = json.loads(json_str)

        ftc = data["ftc_summary"]
        assert ftc is not None
        assert "foreign_tax_paid_total" in ftc
        assert "creditable_total" in ftc
        assert "records" in ftc

    def test_json_write_to_file(self):
        tmp = os.path.join(tempfile.mkdtemp(), "test_cz.json")
        json_str = export_cz_to_json(self.result, output=tmp)
        assert os.path.exists(tmp)
        with open(tmp, "r") as f:
            data = json.load(f)
        assert data["metadata"]["country"] == "cz"

    def test_json_write_to_stringio(self):
        buf = io.StringIO()
        export_cz_to_json(self.result, output=buf)
        buf.seek(0)
        data = json.load(buf)
        assert data["metadata"]["country"] == "cz"

    def test_json_decimal_serialisation(self):
        """Decimals must be strings, not floats."""
        json_str = export_cz_to_json(self.result)
        data = json.loads(json_str)

        # Check a known Decimal field in items
        for item in data["items"]:
            if item.get("gain_loss_eur") is not None:
                # Should be a string representation, not a float
                assert isinstance(item["gain_loss_eur"], str)
                break

    def test_json_netting_present(self):
        json_str = export_cz_to_json(self.result)
        data = json.loads(json_str)
        assert "cz_10_netting" in data
        netting = data["cz_10_netting"]
        assert "securities" in netting
        assert "options" in netting
        assert "combined_net_taxable" in netting


# =========================================================================
# XLSX Tests
# =========================================================================

class TestXlsxExport:
    def setup_method(self):
        self.result = _build_test_result()

    def _export_and_load(self):
        from openpyxl import load_workbook
        buf = io.BytesIO()
        export_cz_to_xlsx(self.result, buf)
        buf.seek(0)
        return load_workbook(buf, read_only=True)

    def test_xlsx_creates_expected_sheets(self):
        wb = self._export_and_load()
        sheet_names = wb.sheetnames
        for expected in ["Summary", "Securities", "Options", "Dividends",
                         "Interest", "WithholdingTax", "PendingReview", "Metadata"]:
            assert expected in sheet_names, f"Missing sheet: {expected}"

    def test_xlsx_securities_has_exempt_item(self):
        wb = self._export_and_load()
        ws = wb["Securities"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        # Find header indices
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        exempt_idx = headers.index("is_exempt")
        assert any(row[exempt_idx] is True or row[exempt_idx] == 1 for row in rows), \
            "Securities sheet should contain at least one exempt item"

    def test_xlsx_options_has_items(self):
        wb = self._export_and_load()
        ws = wb["Options"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) >= 1

    def test_xlsx_dividends_has_ftc_columns(self):
        wb = self._export_and_load()
        ws = wb["Dividends"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for col in ["foreign_tax_paid_czk", "actual_creditable_czk", "non_creditable_czk"]:
            assert col in headers, f"Dividends sheet missing FTC column: {col}"

    def test_xlsx_wht_sheet_has_linked_and_unlinked(self):
        wb = self._export_and_load()
        ws = wb["WithholdingTax"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        linked_idx = headers.index("linked")

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) >= 2  # at least 1 linked + 1 unlinked
        linked_values = [row[linked_idx] for row in rows]
        assert True in linked_values or 1 in linked_values, "Should have linked WHT"
        assert False in linked_values or 0 in linked_values or None in linked_values, \
            "Should have unlinked WHT"

    def test_xlsx_pending_review_sheet(self):
        wb = self._export_and_load()
        ws = wb["PendingReview"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) >= 1

    def test_xlsx_write_to_file(self):
        tmp = os.path.join(tempfile.mkdtemp(), "test_cz.xlsx")
        export_cz_to_xlsx(self.result, tmp)
        assert os.path.exists(tmp)
        assert os.path.getsize(tmp) > 0

    def test_xlsx_summary_has_content(self):
        wb = self._export_and_load()
        ws = wb["Summary"]
        # Should have multiple rows of content
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) >= 5


# =========================================================================
# Edge cases
# =========================================================================

class TestExporterEdgeCases:
    def test_json_empty_result_no_crash(self):
        """Export of empty TaxResult should not crash."""
        result = TaxResult(country_code="cz", tax_year=2025, sections={})
        json_str = export_cz_to_json(result)
        data = json.loads(json_str)
        assert data["metadata"]["country"] == "cz"
        assert len(data["items"]) == 0

    def test_xlsx_empty_result_no_crash(self):
        result = TaxResult(country_code="cz", tax_year=2025, sections={})
        buf = io.BytesIO()
        export_cz_to_xlsx(result, buf)
        buf.seek(0)
        assert buf.read(4) == b"PK\x03\x04"  # valid ZIP (XLSX is a ZIP)

    def test_json_missing_ftc_summary_no_crash(self):
        result = TaxResult(
            country_code="cz", tax_year=2025, sections={},
            country_result={"items": [], "currency": "EUR"},
        )
        json_str = export_cz_to_json(result)
        data = json.loads(json_str)
        assert data["ftc_summary"] is None
