# src/countries/cz/exporters/xlsx_exporter.py
"""
XLSX export for Czech tax results.

Creates an openpyxl workbook with multiple sheets from a ``TaxResult``
with CZ-specific ``country_result`` data.

Sheets:
- Summary — main aggregations, §10 netting, FTC summary
- Securities — SECURITY_DISPOSAL items
- Options — option items
- Dividends — dividend / fund distribution items
- Interest — interest items
- WithholdingTax — all WHT records (linked + unlinked)
- PendingReview — items with tax_review_status != RESOLVED
- Metadata — CZ policy config snapshot
"""
from __future__ import annotations

import logging
from decimal import Decimal
from typing import Any, Dict, List, Optional, Sequence, Union, IO

from src.countries.base import TaxResult

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Column definitions per sheet
# ---------------------------------------------------------------------------

_ITEM_COLUMNS = [
    "item_type", "section", "asset_symbol", "asset_isin", "asset_description",
    "asset_category", "event_date", "acquisition_date", "holding_period_days",
    "quantity", "original_amount", "original_currency",
    "amount_eur", "amount_czk",
    "cost_basis_eur", "proceeds_eur", "gain_loss_eur",
    "cost_basis_czk", "proceeds_czk", "gain_loss_czk",
    "is_taxable", "is_exempt", "exemption_reason", "included_in_tax_base",
    "qualifies_for_annual_limit", "exempt_due_to_annual_limit",
    "tax_review_status", "tax_review_note",
    "wht_total_czk",
    "source_country",
    "fx_source", "fx_policy", "fx_rate", "fx_date_used",
]

_FTC_COLUMNS = [
    "foreign_tax_paid_czk", "configured_credit_cap_rate",
    "max_creditable_czk", "actual_creditable_czk", "non_creditable_czk",
    "ftc_review_status", "ftc_review_note",
]

_WHT_COLUMNS = [
    "parent_item_type", "parent_event_id", "parent_asset_symbol",
    "wht_event_id", "event_date", "original_amount", "original_currency",
    "amount_czk", "source_country", "fx_rate", "fx_date_used",
    "linked",
]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def export_cz_to_xlsx(
    tax_result: TaxResult,
    output: Union[str, IO[bytes]],
) -> None:
    """
    Export a CZ ``TaxResult`` to an XLSX workbook.

    Args:
        tax_result: The ``TaxResult`` from ``CzechTaxAggregator.aggregate()``.
        output: File path (str) or file-like object to write to.
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    cr = tax_result.country_result or {}
    items: list = cr.get("items", [])
    netting = cr.get("netting")
    ftc_summary = cr.get("ftc_summary")
    cur = cr.get("currency", "EUR")

    # Helper to convert item to flat row dict (including FTC)
    def _item_row(it: Any) -> Dict[str, Any]:
        d = it.to_dict()
        # Source country from first WHT record
        if "source_country" not in d or d.get("source_country") is None:
            for r in getattr(it, "wht_records", []):
                if r.source_country:
                    d["source_country"] = r.source_country
                    break
        # FTC fields
        ftc_rec = getattr(it, "ftc_record", None)
        if ftc_rec is not None:
            d["foreign_tax_paid_czk"] = str(ftc_rec.foreign_tax_paid_czk)
            d["configured_credit_cap_rate"] = str(ftc_rec.configured_cap_rate)
            d["max_creditable_czk"] = str(ftc_rec.max_creditable_czk)
            d["actual_creditable_czk"] = str(ftc_rec.actual_creditable_czk)
            d["non_creditable_czk"] = str(ftc_rec.non_creditable_czk)
            d["ftc_review_status"] = ftc_rec.review_status
            d["ftc_review_note"] = ftc_rec.review_note
        return d

    # --- Sheet: Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary_sheet(ws_summary, tax_result, netting, ftc_summary, cur)

    # --- Sheet: Securities ---
    from src.countries.cz.tax_items import CzTaxItemType
    sec_items = [it for it in items if it.item_type == CzTaxItemType.SECURITY_DISPOSAL]
    ws_sec = wb.create_sheet("Securities")
    _write_items_sheet(ws_sec, sec_items, _ITEM_COLUMNS, _item_row)

    # --- Sheet: Options ---
    opt_items = [it for it in items if it.item_type in (
        CzTaxItemType.OPTION_CLOSE, CzTaxItemType.OPTION_EXPIRY_WORTHLESS,
        CzTaxItemType.OPTION_EXERCISE_ASSIGNMENT,
    )]
    ws_opt = wb.create_sheet("Options")
    _write_items_sheet(ws_opt, opt_items, _ITEM_COLUMNS, _item_row)

    # --- Sheet: Dividends ---
    div_items = [it for it in items if it.item_type in (
        CzTaxItemType.DIVIDEND, CzTaxItemType.FUND_DISTRIBUTION,
    )]
    ws_div = wb.create_sheet("Dividends")
    cols = _ITEM_COLUMNS + _FTC_COLUMNS
    _write_items_sheet(ws_div, div_items, cols, _item_row)

    # --- Sheet: Interest ---
    int_items = [it for it in items if it.item_type == CzTaxItemType.INTEREST]
    ws_int = wb.create_sheet("Interest")
    _write_items_sheet(ws_int, int_items, _ITEM_COLUMNS + _FTC_COLUMNS, _item_row)

    # --- Sheet: WithholdingTax ---
    ws_wht = wb.create_sheet("WithholdingTax")
    _write_wht_sheet(ws_wht, items)

    # --- Sheet: PendingReview ---
    from src.countries.cz.tax_items import CzTaxReviewStatus
    pending = [it for it in items if it.tax_review_status == CzTaxReviewStatus.PENDING_MANUAL_REVIEW]
    ws_pending = wb.create_sheet("PendingReview")
    _write_items_sheet(ws_pending, pending, _ITEM_COLUMNS, _item_row)

    # --- Sheet: Metadata ---
    ws_meta = wb.create_sheet("Metadata")
    _write_metadata_sheet(ws_meta, tax_result, cr)

    # Freeze header rows on all item sheets
    for ws in [ws_sec, ws_opt, ws_div, ws_int, ws_wht, ws_pending]:
        ws.freeze_panes = "A2"

    # Save
    if isinstance(output, str):
        wb.save(output)
        logger.info(f"CZ XLSX export written to {output}")
    else:
        wb.save(output)


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_items_sheet(
    ws: Any,
    items: list,
    columns: List[str],
    row_fn: Any,
) -> None:
    """Write a header row + one row per item."""
    # Header
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Rows
    for row_idx, it in enumerate(items, 2):
        d = row_fn(it)
        for col_idx, col_name in enumerate(columns, 1):
            val = d.get(col_name)
            # Convert string-Decimals back for Excel
            if isinstance(val, str):
                try:
                    val = float(Decimal(val))
                except Exception:
                    pass
            if isinstance(val, Decimal):
                val = float(val)
            ws.cell(row=row_idx, column=col_idx, value=val)


def _write_wht_sheet(ws: Any, items: list) -> None:
    """All WHT records: linked (from income items) + unlinked (OTHER items)."""
    from src.countries.cz.tax_items import CzTaxItemType

    # Header
    for col_idx, col_name in enumerate(_WHT_COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    row_idx = 2
    for it in items:
        for r in it.wht_records:
            is_linked = it.item_type != CzTaxItemType.OTHER
            vals = {
                "parent_item_type": it.item_type.name,
                "parent_event_id": str(it.source_event_id),
                "parent_asset_symbol": it.asset_symbol,
                "wht_event_id": str(r.wht_event_id),
                "event_date": r.event_date,
                "original_amount": float(r.original_amount) if r.original_amount else None,
                "original_currency": r.original_currency,
                "amount_czk": float(r.amount_czk) if r.amount_czk else None,
                "source_country": r.source_country,
                "fx_rate": float(r.fx.fx_rate_inverse) if r.fx else None,
                "fx_date_used": r.fx.fx_date_used if r.fx else None,
                "linked": is_linked,
            }
            for col_idx, col_name in enumerate(_WHT_COLUMNS, 1):
                ws.cell(row=row_idx, column=col_idx, value=vals.get(col_name))
            row_idx += 1


def _write_summary_sheet(
    ws: Any,
    tax_result: TaxResult,
    netting: Any,
    ftc_summary: Any,
    currency: str,
) -> None:
    """Summary sheet with section aggregations."""
    row = 1

    def _header(text: str) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=text)
        row += 1

    def _kv(key: str, value: Any) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=key)
        v = float(value) if isinstance(value, Decimal) else value
        ws.cell(row=row, column=2, value=v)
        row += 1

    _header(f"Czech Tax Summary — {tax_result.tax_year} ({currency})")
    row += 1

    # Sections
    for sec_key, section in tax_result.sections.items():
        _header(f"— {section.label} —")
        for k, v in section.line_items.items():
            _kv(k, v)
        for note in section.notes:
            _kv("⚠ NOTE", note)
        row += 1


def _write_metadata_sheet(ws: Any, tax_result: TaxResult, cr: Dict) -> None:
    """Config snapshot for audit trail."""
    row = 1
    ws.cell(row=row, column=1, value="Key")
    ws.cell(row=row, column=2, value="Value")
    row += 1

    def _kv(key: str, value: Any) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=str(value))
        row += 1

    _kv("country", tax_result.country_code)
    _kv("tax_year", tax_result.tax_year)
    _kv("currency", cr.get("currency", "EUR"))

    fx_policy = cr.get("fx_policy")
    if fx_policy:
        _kv("fx_mode", fx_policy.mode.name)
        _kv("fx_source", fx_policy.source)
        _kv("fx_weekend_fallback", fx_policy.weekend_fallback.name)

    _kv("fx_records_count", len(cr.get("fx_conversion_records", [])))
    _kv("total_items", len(cr.get("items", [])))
